import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook

# Configurações
NOME_ARQUIVO = 'Sign Off.xlsx'
CAMINHO_ALTERNATIVO = r'C:\Pasta Main\Excel\Sign Off.xlsx'  # Caminho alternativo, pode estar vazio
CAMINHO_PADRAO = r'C:\Pasta Main\Excel\Sign Off.xlsx'

COLUNA_IDH = 6     # Coluna G (índice 6)
COLUNA_UM = 7      # Coluna H (índice 7)
COLUNA_INICIO_VALORES = 8   # Coluna I (índice 8)
COLUNA_FIM_VALORES = 80     # Coluna CB (índice 80)
ABA_PADRAO = 'Standart Unit'
ABA_CONVERSAO = 'Conversão'

def localizar_arquivo():
    if os.path.exists(NOME_ARQUIVO):
        print(f"✔ Arquivo encontrado no diretório atual.")
        return os.path.abspath(NOME_ARQUIVO), os.path.dirname(os.path.abspath(NOME_ARQUIVO))
    else:
        print("❌ Arquivo não encontrado no diretório atual.")

        if CAMINHO_ALTERNATIVO:
            if os.path.exists(CAMINHO_ALTERNATIVO):
                print(f"✔ Utilizando caminho alternativo: {CAMINHO_ALTERNATIVO}")
                return CAMINHO_ALTERNATIVO, os.path.dirname(CAMINHO_ALTERNATIVO)
            else:
                raise FileNotFoundError("Arquivo não encontrado no caminho alternativo especificado.")
        else:
            caminho_padrao_completo = os.path.join(CAMINHO_PADRAO, NOME_ARQUIVO)
            if os.path.exists(caminho_padrao_completo):
                print(f"✔ Utilizando caminho padrão: {caminho_padrao_completo}")
                return caminho_padrao_completo, CAMINHO_PADRAO
            else:
                raise FileNotFoundError("Arquivo não encontrado nem no diretório atual, nem no caminho alternativo, nem no caminho padrão.")

def deletar_arquivos_antigos(pasta, base_nome="Sign Off - "):
    agora = datetime.now()
    for nome_arquivo in os.listdir(pasta):
        if nome_arquivo.startswith(base_nome) and nome_arquivo.endswith(".xlsx"):
            match = re.search(r'Sign Off - (\d{4})-(\d{2})-(\d{2})', nome_arquivo)
            if match:
                ano, mes, dia = map(int, match.groups())
                try:
                    data_arquivo = datetime(ano, mes, dia)
                    delta = agora - data_arquivo
                    if delta.days > 0 and (delta.days > 31 or data_arquivo.month < agora.month or data_arquivo.year < agora.year):
                        caminho_completo = os.path.join(pasta, nome_arquivo)
                        os.remove(caminho_completo)
                        print(f" Arquivo antigo removido: {nome_arquivo}")
                except ValueError:
                    continue

try:
    caminho_arquivo, pasta_destino = localizar_arquivo()

    # Usamos pandas só para a lógica de comparação
    df_conversao = pd.read_excel(caminho_arquivo, sheet_name=ABA_CONVERSAO)
    if df_conversao.shape[1] < 2:
        raise ValueError("A aba 'Conversão' precisa conter pelo menos duas colunas: IDH (coluna A) e Peso Kg (coluna B).")
    conversao_dict = dict(zip(df_conversao.iloc[:, 0], df_conversao.iloc[:, 1]))

    df_standart = pd.read_excel(caminho_arquivo, sheet_name=ABA_PADRAO)
    indices_para_converter = df_standart[df_standart.iloc[:, COLUNA_UM].str.upper() != "KG"].index

    # Carregando o arquivo com openpyxl para edição com formatação preservada
    wb = load_workbook(caminho_arquivo)
    ws = wb[ABA_PADRAO]

    for i in indices_para_converter:
        idh = df_standart.iat[i, COLUNA_IDH]
        fator = conversao_dict.get(idh, None)
        if fator is None:
            continue
        for col in range(COLUNA_INICIO_VALORES, COLUNA_FIM_VALORES + 1):
            cell = ws.cell(row=i + 2, column=col + 1)  # +2 pois openpyxl começa do 1 e há header
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value * fator

    deletar_arquivos_antigos(pasta_destino)

    data_execucao = datetime.now().strftime('%Y-%m-%d-%H%M')
    novo_nome = f"Sign Off - {data_execucao}.xlsx"
    caminho_salvar = os.path.join(pasta_destino, novo_nome)

    wb.save(caminho_salvar)
    print(f"✔ Arquivo salvo com o nome: {caminho_salvar}")

except Exception as e:
    print(f"\n🚫 ERRO: {e}")
    print("O script encontrou um problema durante a execução.")
